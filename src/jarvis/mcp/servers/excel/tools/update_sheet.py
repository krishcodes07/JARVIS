"""
Update Sheet Tool for Excel.
"""

import os
import openpyxl
from ..config import safe_excel_path

NAME = "update_sheet"
DESCRIPTION = (
    "Update specific cells in an Excel worksheet by cell address (e.g. 'A1', 'B3') "
    "or by row/column index. Supports single-cell and batch updates."
)


def update_sheet(
    filename: str,
    updates: list[dict],
    sheet_name: str | None = None,
) -> str:
    """
    Update specific cells in an Excel worksheet.

    Each update entry must contain a cell reference and a value. Supported formats:

    - By cell address:  {"cell": "B2", "value": 42}
    - By row/col index: {"row": 2, "col": 3, "value": "Hello"}  (1-indexed)

    Args:
        filename: Name of the Excel file (must already exist).
        updates: List of update dicts, each specifying a cell and new value.
        sheet_name: Worksheet name (default: active sheet).

    Returns:
        Success message with count of updated cells, or an error message.

    Examples:
        update_sheet("budget.xlsx", [{"cell": "A1", "value": "Revenue"}])
        update_sheet("data.xlsx", [{"row": 2, "col": 3, "value": 99.5}], sheet_name="Q1")
    """
    try:
        filepath = safe_excel_path(filename)
        if not os.path.exists(filepath):
            return f"❌ File not found: {filename}"

        wb = openpyxl.load_workbook(filepath)

        if sheet_name:
            if sheet_name not in wb.sheetnames:
                return f"❌ Sheet '{sheet_name}' not found in '{filename}'. Available: {wb.sheetnames}"
            ws = wb[sheet_name]
        else:
            ws = wb.active
            if ws is None:
                return f"❌ No active sheet found in '{filename}'. Please specify a sheet_name."

        if not updates:
            return "⚠️ No updates provided."

        updated = 0
        errors = []

        for i, entry in enumerate(updates):
            value = entry.get("value")
            try:
                if "cell" in entry:
                    ws[entry["cell"]] = value
                    updated += 1
                elif "row" in entry and "col" in entry:
                    row = int(entry["row"])
                    col = int(entry["col"])
                    ws.cell(row=row, column=col, value=value)
                    updated += 1
                else:
                    errors.append(
                        f"Entry {i}: missing 'cell' or 'row'/'col' keys — {entry}"
                    )
            except Exception as cell_err:
                errors.append(f"Entry {i}: {cell_err}")

        wb.save(filepath)

        sheet_title = ws.title if ws is not None else "(unknown)"
        msg = f"✅ Updated {updated} cell(s) in '{filename}' sheet '{sheet_title}'."
        if errors:
            msg += "\n⚠️ Errors:\n" + "\n".join(f"  - {e}" for e in errors)
        return msg

    except Exception as e:
        return f"❌ Error updating sheet: {e}"
