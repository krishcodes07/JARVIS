"""
Read sheet tool for Excel.
"""

import os
import openpyxl
from ..config import safe_excel_path

NAME = "read_sheet"
DESCRIPTION = "Read data from an Excel worksheet."


def read_sheet(
    filename: str,
    sheet_name: str | None = None,
    max_rows: int = 50,
) -> str:
    """
    Read data from an Excel worksheet.

    Args:
        filename: Name of the Excel file.
        sheet_name: Worksheet name (default: active sheet).
        max_rows: Maximum rows to read (default: 50).

    Returns:
        Formatted text representation of the worksheet content.
    """
    try:
        filepath = safe_excel_path(filename)
        if not os.path.exists(filepath):
            return f"❌ File not found: {filename}"

        wb = openpyxl.load_workbook(filepath, data_only=True)

        if sheet_name and sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
        else:
            ws = wb.active
            if ws is None:
                return f"❌ No active sheet found in '{filename}'. Please specify a sheet_name."

        rows = list(ws.iter_rows(values_only=True))
        sheet_title = ws.title if ws is not None else "(unknown)"
        if not rows:
            return f"📊 Sheet '{sheet_title}' in '{filename}' is empty."

        lines = [f"📊 Content of '{filename}' → sheet '{sheet_title}':"]
        for i, row in enumerate(rows[:max_rows], 1):
            row_str = " | ".join(str(val) if val is not None else "" for val in row)
            lines.append(f"  Row {i:2d}: {row_str}")

        if len(rows) > max_rows:
            lines.append(f"  ... (truncated, showing {max_rows} of {len(rows)} rows)")

        return "\n".join(lines)

    except Exception as e:
        return f"❌ Error reading sheet: {e}"
