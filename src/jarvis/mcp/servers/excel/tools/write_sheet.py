"""
Write sheet tool for Excel.
"""

import os

import openpyxl

from ..config import safe_excel_path

NAME = "write_sheet"
DESCRIPTION = "Append rows or write data to a specific sheet in an Excel workbook."


def write_sheet(
    filename: str,
    rows: list[list[str | int | float]],
    sheet_name: str | None = None,
    mode: str = "append",
) -> str:
    """
    Write data to an Excel worksheet.

    Args:
        filename: Name of the Excel file.
        rows: List of rows to write, where each row is a list of cell values.
        sheet_name: Sheet name (default: active sheet).
        mode: "append" to add rows at bottom, "overwrite" to clear and write.

    Returns:
        Success or error message.
    """
    try:
        filepath = safe_excel_path(filename)

        if os.path.exists(filepath):
            wb = openpyxl.load_workbook(filepath)
        else:
            wb = openpyxl.Workbook()

        if sheet_name:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
            else:
                ws = wb.create_sheet(title=sheet_name)
        else:
            ws = wb.active
            if ws is None:
                return f"❌ No active sheet found in '{filename}'. Please specify a sheet_name."

        if mode == "overwrite":
            ws.delete_rows(1, ws.max_row + 1)

        added = 0
        for row_data in rows:
            ws.append(row_data)
            added += 1

        wb.save(filepath)
        sheet_title = ws.title if ws is not None else "(unknown)"
        return f"✅ Wrote {added} rows to '{filename}' sheet '{sheet_title}' (mode: {mode})"

    except Exception as e:
        return f"❌ Error writing sheet: {e}"
